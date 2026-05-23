from __future__ import annotations

import csv
import json
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Literal, Any

from cbz_manga_translator.core.cache import ProjectCache
from cbz_manga_translator.core.models import ProjectData, OcrBlock

RiskBand = Literal["high", "medium", "ok"]

_DECISION_VALUES = ["", "validate", "correct", "review", "fused", "ignore", "sfx"]

# User-facing order: each editable field is placed immediately next to the
# evidence needed to fill it. Technical identifiers are pushed to the right.
_REVIEW_FIELDNAMES = [
    "review_decision",
    "suggested_action",
    "risk_score",
    "risk_reasons",
    "quality_warnings",
    "review_notes",
    "ocr_text",
    "corrected_ocr",
    "ocr_corrected_text",
    "source_for_review",
    "corrected_source",
    "normalized_source_text",
    "translation_fr",
    "corrected_fr",
    "raw_translation_fr",
    "series_label",
    "volume_label",
    "page_number",
    "image_name",
    "block_id",
    "page_index",
    "bbox",
    "source_lang",
    "confidence",
    "ocr_alternatives_count",
    "ocr_alternatives",
]

_EDITABLE_FIELDS = {
    "review_decision",
    "corrected_ocr",
    "corrected_source",
    "corrected_fr",
    "review_notes",
}

_HEADER_LABELS = {
    "review_decision": "Décision",
    "suggested_action": "Action suggérée",
    "risk_score": "Risque",
    "risk_reasons": "Raisons risque",
    "quality_warnings": "Warnings QC",
    "review_notes": "Notes reviewer",
    "ocr_text": "OCR brut à lire",
    "corrected_ocr": "OCR corrigé à remplir",
    "ocr_corrected_text": "OCR corrigé auto",
    "source_for_review": "Source à vérifier",
    "corrected_source": "Source corrigée à remplir",
    "normalized_source_text": "Source normalisée auto",
    "translation_fr": "Traduction actuelle",
    "corrected_fr": "Traduction FR corrigée à remplir",
    "raw_translation_fr": "Trad brute",
    "series_label": "Série",
    "volume_label": "Tome",
    "page_number": "Page",
    "image_name": "Image",
    "block_id": "Block ID",
    "page_index": "Page index",
    "bbox": "BBox",
    "source_lang": "Langue source",
    "confidence": "Confiance OCR",
    "ocr_alternatives_count": "Nb alternatives OCR",
    "ocr_alternatives": "Alternatives OCR",
}

_DECISION_HELP = """# MangaTrad human review pack

Le fichier recommandé est maintenant `mangatrad_human_review_pack.xlsx`.
Il est plus simple à corriger que le TSV : colonnes lisibles, filtres, lignes figées et liste déroulante dans `Décision`.

Colonnes à remplir uniquement :

- `Décision` : liste déroulante `validate`, `correct`, `review`, `fused`, `ignore`, `sfx`.
- `OCR corrigé à remplir` : OCR corrigé si l'OCR brut est faux.
- `Source corrigée à remplir` : texte source final, si la normalisation doit être forcée.
- `Traduction FR corrigée à remplir` : traduction française corrigée.
- `Notes reviewer` : commentaire libre.

Les colonnes de contexte sont placées juste à côté des champs à remplir :

- `OCR brut à lire` → `OCR corrigé à remplir`
- `Source à vérifier` → `Source corrigée à remplir`
- `Traduction actuelle` → `Traduction FR corrigée à remplir`

Décisions :

- `validate` : la traduction finale est correcte telle quelle.
- `correct` : tu as rempli au moins une correction OCR/source/FR.
- `review` : bloc à revoir plus tard.
- `fused` : plusieurs bulles, ou SFX mélangé à une bulle ; à retraiter/séparer.
- `ignore` : bloc inutile pour l'apprentissage.
- `sfx` : bruit/SFX/onomatopée ou texte non-dialogue à ignorer.

Ne modifie pas les colonnes techniques à droite : `page_index`, `block_id`, `image_name`, `bbox`.
Le TSV reste généré uniquement comme fallback/compatibilité.
"""


@dataclass(slots=True)
class ReviewPackResult:
    review_path: Path
    workbook_path: Path
    jsonl_path: Path
    guide_path: Path
    selected_rows: int
    total_rows: int
    series_count: int


@dataclass(slots=True)
class ReviewApplyResult:
    output_project_path: Path
    changed_blocks: int
    validated_blocks: int
    corrected_blocks: int
    ignored_blocks: int
    review_blocks: int
    skipped_rows: int


def _read_analysis_rows(analysis_dir: str | Path) -> list[dict[str, str]]:
    analysis_path = Path(analysis_dir)
    csv_path = analysis_path / "mangatrad_review_blocks.csv"
    if not csv_path.exists():
        raise FileNotFoundError(f"CSV d'analyse introuvable: {csv_path}")
    with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [dict(row) for row in csv.DictReader(handle)]


def _risk_band(row: dict[str, str]) -> RiskBand:
    try:
        score = int(float(row.get("risk_score") or 0))
    except ValueError:
        score = 0
    if score >= 55:
        return "high"
    if score >= 25:
        return "medium"
    return "ok"


def _normalized_series(row: dict[str, str]) -> str:
    return (row.get("series_label") or "unknown").strip() or "unknown"


def _row_sort_key(row: dict[str, str]) -> tuple[int, int, str]:
    try:
        risk = int(float(row.get("risk_score") or 0))
    except ValueError:
        risk = 0
    try:
        page_index = int(row.get("page_index") or 0)
    except ValueError:
        page_index = 0
    return (-risk, page_index, row.get("block_id") or "")


def _round_robin_by_series(rows: Iterable[dict[str, str]], limit: int | None) -> list[dict[str, str]]:
    buckets: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        buckets[_normalized_series(row)].append(row)
    for bucket in buckets.values():
        bucket.sort(key=_row_sort_key)
    selected: list[dict[str, str]] = []
    series_order = sorted(buckets, key=lambda series: (-len(buckets[series]), series.lower()))
    while series_order and (limit is None or len(selected) < limit):
        next_order: list[str] = []
        for series in series_order:
            bucket = buckets[series]
            if bucket:
                selected.append(bucket.pop(0))
                if limit is not None and len(selected) >= limit:
                    break
            if bucket:
                next_order.append(series)
        series_order = next_order
    return selected


def _prepare_review_row(row: dict[str, str]) -> dict[str, str]:
    prepared = {field: str(row.get(field, "")) for field in _REVIEW_FIELDNAMES}
    prepared["review_decision"] = ""
    prepared["corrected_ocr"] = ""
    prepared["corrected_source"] = ""
    prepared["corrected_fr"] = ""
    prepared["review_notes"] = ""
    return prepared


def _write_review_tsv(path: Path, rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=_REVIEW_FIELDNAMES,
            delimiter="\t",
            quoting=csv.QUOTE_ALL,
        )
        writer.writeheader()
        writer.writerows(rows)


def _write_review_workbook(path: Path, rows: list[dict[str, str]]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.comments import Comment
        from openpyxl.formatting.rule import CellIsRule, FormulaRule
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ModuleNotFoundError as exc:  # pragma: no cover - exercised in user env if dependency missing
        raise RuntimeError(
            "openpyxl est requis pour générer le pack XLSX. Installe-le avec `python -m pip install openpyxl`."
        ) from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "Review"
    choices = wb.create_sheet("Choices")
    choices.append(["review_decision"])
    for value in _DECISION_VALUES:
        choices.append([value])
    choices.sheet_state = "hidden"

    header = [_HEADER_LABELS.get(field, field) for field in _REVIEW_FIELDNAMES]
    ws.append(header)
    for row in rows:
        ws.append([row.get(field, "") for field in _REVIEW_FIELDNAMES])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    header_fill = PatternFill("solid", fgColor="1F2937")
    editable_fill = PatternFill("solid", fgColor="FEF3C7")
    context_fill = PatternFill("solid", fgColor="E0F2FE")
    tech_fill = PatternFill("solid", fgColor="F3F4F6")
    high_fill = PatternFill("solid", fgColor="FEE2E2")
    medium_fill = PatternFill("solid", fgColor="FEF3C7")
    ok_fill = PatternFill("solid", fgColor="DCFCE7")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, field in enumerate(_REVIEW_FIELDNAMES, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        if field in _EDITABLE_FIELDS:
            cell.comment = Comment("Colonne à remplir/modifier par le reviewer.", "MangaTrad")
        elif field in {"page_index", "block_id", "bbox", "image_name"}:
            cell.comment = Comment("Colonne technique : ne pas modifier.", "MangaTrad")

    widths = {
        "review_decision": 15,
        "suggested_action": 18,
        "risk_score": 9,
        "risk_reasons": 34,
        "quality_warnings": 34,
        "review_notes": 30,
        "ocr_text": 38,
        "corrected_ocr": 38,
        "ocr_corrected_text": 34,
        "source_for_review": 42,
        "corrected_source": 42,
        "normalized_source_text": 38,
        "translation_fr": 42,
        "corrected_fr": 46,
        "raw_translation_fr": 34,
        "series_label": 24,
        "volume_label": 24,
        "page_number": 9,
        "image_name": 36,
        "block_id": 18,
        "page_index": 10,
        "bbox": 18,
        "source_lang": 10,
        "confidence": 10,
        "ocr_alternatives_count": 12,
        "ocr_alternatives": 45,
    }
    for col_idx, field in enumerate(_REVIEW_FIELDNAMES, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = widths.get(field, 18)
        fill = editable_fill if field in _EDITABLE_FIELDS else tech_fill
        if field in {
            "suggested_action", "risk_score", "risk_reasons", "quality_warnings",
            "ocr_text", "ocr_corrected_text", "source_for_review", "normalized_source_text",
            "translation_fr", "raw_translation_fr",
        }:
            fill = context_fill
        for row_idx in range(2, len(rows) + 2):
            c = ws.cell(row=row_idx, column=col_idx)
            c.fill = fill
            c.border = border
            c.alignment = Alignment(vertical="top", wrap_text=True)

    # Decision dropdown.
    decision_col = _REVIEW_FIELDNAMES.index("review_decision") + 1
    decision_letter = get_column_letter(decision_col)
    dv = DataValidation(type="list", formula1='"validate,correct,review,fused,ignore,sfx"', allow_blank=True)
    dv.error = "Choisis validate, correct, review, fused, ignore ou sfx."
    dv.errorTitle = "Décision invalide"
    dv.prompt = "validate = OK, correct = correction faite, review = à revoir, fused = bulles/SFX fusionnés, ignore = inutile, sfx = bruit/onomatopée."
    dv.promptTitle = "Décision MangaTrad"
    ws.add_data_validation(dv)
    dv.add(f"{decision_letter}2:{decision_letter}{len(rows)+1}")

    risk_col = _REVIEW_FIELDNAMES.index("risk_score") + 1
    risk_letter = get_column_letter(risk_col)
    if rows:
        ws.conditional_formatting.add(
            f"{risk_letter}2:{risk_letter}{len(rows)+1}",
            CellIsRule(operator="greaterThanOrEqual", formula=["55"], fill=high_fill),
        )
        ws.conditional_formatting.add(
            f"{risk_letter}2:{risk_letter}{len(rows)+1}",
            CellIsRule(operator="between", formula=["25", "54"], fill=medium_fill),
        )
        ws.conditional_formatting.add(
            f"{risk_letter}2:{risk_letter}{len(rows)+1}",
            CellIsRule(operator="lessThan", formula=["25"], fill=ok_fill),
        )
        # Highlight rows that still lack a decision.
        ws.conditional_formatting.add(
            f"A2:{get_column_letter(len(_REVIEW_FIELDNAMES))}{len(rows)+1}",
            FormulaRule(formula=[f'$A2=""'], fill=PatternFill("solid", fgColor="FFF7ED")),
        )

    for row_idx in range(2, len(rows) + 2):
        ws.row_dimensions[row_idx].height = 72
    ws.row_dimensions[1].height = 34

    instructions = wb.create_sheet("Instructions")
    instructions["A1"] = "MangaTrad — correction humaine"
    instructions["A1"].font = Font(bold=True, size=15)
    instructions["A3"] = "1. Remplis surtout les cellules jaunes."
    instructions["A4"] = "2. Utilise la liste déroulante Décision."
    instructions["A5"] = "3. OCR brut → OCR corrigé ; Source à vérifier → Source corrigée ; Traduction actuelle → Traduction corrigée."
    instructions["A6"] = "4. Ne modifie pas les colonnes techniques à droite : page_index, block_id, image_name, bbox."
    instructions["A7"] = "5. Mets fused quand la bbox mélange plusieurs bulles/SFX ; sauvegarde le .xlsx, puis réinjecte-le avec corpus_apply_review."
    instructions.column_dimensions["A"].width = 120

    wb.save(path)


def create_review_pack(
    analysis_dir: str | Path,
    output_dir: str | Path,
    *,
    max_blocks: int = 200,
    include_high: bool = True,
    include_medium: bool = True,
    include_ok: bool = False,
    balanced: bool = True,
) -> ReviewPackResult:
    rows = _read_analysis_rows(analysis_dir)
    allowed_bands = set()
    if include_high:
        allowed_bands.add("high")
    if include_medium:
        allowed_bands.add("medium")
    if include_ok:
        allowed_bands.add("ok")
    candidates = [row for row in rows if _risk_band(row) in allowed_bands]
    if balanced:
        selected = _round_robin_by_series(candidates, max_blocks)
    else:
        selected = sorted(candidates, key=_row_sort_key)[:max_blocks]
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    review_path = output_path / "mangatrad_human_review_pack.tsv"
    workbook_path = output_path / "mangatrad_human_review_pack.xlsx"
    jsonl_path = output_path / "mangatrad_human_review_pack.jsonl"
    guide_path = output_path / "mangatrad_human_review_guide.md"
    prepared_rows = [_prepare_review_row(row) for row in selected]
    _write_review_tsv(review_path, prepared_rows)
    _write_review_workbook(workbook_path, prepared_rows)
    with jsonl_path.open("w", encoding="utf-8") as handle:
        for row in prepared_rows:
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")
    guide_path.write_text(_DECISION_HELP, encoding="utf-8")
    return ReviewPackResult(
        review_path=review_path,
        workbook_path=workbook_path,
        jsonl_path=jsonl_path,
        guide_path=guide_path,
        selected_rows=len(prepared_rows),
        total_rows=len(rows),
        series_count=len({_normalized_series(row) for row in selected}),
    )


def _find_block_index(project: ProjectData) -> dict[tuple[int, str], OcrBlock]:
    index: dict[tuple[int, str], OcrBlock] = {}
    for page in project.pages:
        for block in page.blocks:
            index[(page.page_index, block.id)] = block
    return index


def _clean_cell(value: Any | None) -> str:
    return str(value or "").strip()


def _decision(value: Any | None) -> str:
    return _clean_cell(value).lower().replace("à", "a")


def _detect_review_delimiter(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".tsv":
        return "\t"
    if suffix == ".csv":
        return ","
    try:
        sample = path.read_text(encoding="utf-8-sig", errors="replace")[:4096]
    except OSError:
        return "\t"
    first = sample.splitlines()[0] if sample.splitlines() else ""
    if "\t" in first:
        return "\t"
    try:
        return csv.Sniffer().sniff(sample, delimiters="\t,;").delimiter
    except csv.Error:
        return "\t"


def _read_xlsx_review_rows(path: Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:  # pragma: no cover
        raise RuntimeError(
            "openpyxl est requis pour lire le pack XLSX. Installe-le avec `python -m pip install openpyxl`."
        ) from exc
    wb = load_workbook(path, data_only=True, read_only=True)
    if "Review" not in wb.sheetnames:
        raise ValueError(f"Feuille 'Review' introuvable dans {path}")
    ws = wb["Review"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    reverse_labels = {label: field for field, label in _HEADER_LABELS.items()}
    headers = [reverse_labels.get(str(cell or "").strip(), str(cell or "").strip()) for cell in rows[0]]
    result: list[dict[str, str]] = []
    for values in rows[1:]:
        row = {headers[idx]: _clean_cell(value) for idx, value in enumerate(values) if idx < len(headers)}
        if any(row.values()):
            result.append(row)
    return result


def _read_review_rows(review_table: str | Path) -> list[dict[str, str]]:
    path = Path(review_table)
    if path.suffix.lower() == ".xlsx":
        return _read_xlsx_review_rows(path)
    delimiter = _detect_review_delimiter(path)
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [dict(row) for row in csv.DictReader(handle, delimiter=delimiter)]


def apply_review_pack(
    project_path: str | Path,
    review_csv: str | Path,
    *,
    output_project_path: str | Path | None = None,
) -> ReviewApplyResult:
    project = ProjectCache.load(project_path)
    block_index = _find_block_index(project)
    rows = _read_review_rows(review_csv)
    changed = 0
    validated = 0
    corrected = 0
    ignored = 0
    review = 0
    skipped = 0
    for row in rows:
        try:
            page_index = int(row.get("page_index") or -1)
        except ValueError:
            skipped += 1
            continue
        block_id = _clean_cell(row.get("block_id"))
        block = block_index.get((page_index, block_id))
        if block is None:
            skipped += 1
            continue
        decision = _decision(row.get("review_decision"))
        corrected_ocr = _clean_cell(row.get("corrected_ocr"))
        corrected_source = _clean_cell(row.get("corrected_source"))
        corrected_fr = _clean_cell(row.get("corrected_fr"))
        review_notes = _clean_cell(row.get("review_notes"))
        row_changed = False
        if corrected_ocr and corrected_ocr != block.ocr_corrected_text:
            block.ocr_corrected_text = corrected_ocr
            row_changed = True
        if corrected_source and corrected_source != block.normalized_source_text:
            block.normalized_source_text = corrected_source
            row_changed = True
        if corrected_fr and corrected_fr != block.translation_fr:
            block.translation_fr = corrected_fr
            block.raw_translation_fr = block.raw_translation_fr or corrected_fr
            row_changed = True
        if decision in {"sfx", "noise"} and not review_notes.lower().startswith("[sfx]"):
            review_notes = f"[sfx] {review_notes}".strip()
        if decision in {"fused", "fusion", "merged", "merge"} and not review_notes.lower().startswith("[fusion]"):
            review_notes = f"[fusion] {review_notes}".strip()
        if review_notes and getattr(block, "review_notes", "") != review_notes:
            block.review_notes = review_notes
            row_changed = True
        if decision in {"ignore", "ignored", "sfx", "noise"}:
            block.manual_status = "ignored"
            ignored += 1
            row_changed = True
        elif decision in {"review", "revoir", "a revoir", "todo", "fused", "fusion", "merged", "merge"}:
            block.manual_status = "review"
            review += 1
            row_changed = True
        elif decision in {"validate", "valid", "validated", "ok", "keep"}:
            block.manual_status = "validated"
            validated += 1
            row_changed = True
        elif decision in {"correct", "corrected", "edit", "edited"} or corrected_ocr or corrected_source or corrected_fr:
            block.manual_status = "edited"
            corrected += 1
            row_changed = True
        elif not decision:
            skipped += 1
            continue
        else:
            block.manual_status = "review"
            review += 1
            row_changed = True
        if row_changed:
            changed += 1
    output_path = Path(output_project_path) if output_project_path else Path(project_path)
    ProjectCache.save(output_path, project)
    return ReviewApplyResult(
        output_project_path=output_path,
        changed_blocks=changed,
        validated_blocks=validated,
        corrected_blocks=corrected,
        ignored_blocks=ignored,
        review_blocks=review,
        skipped_rows=skipped,
    )
